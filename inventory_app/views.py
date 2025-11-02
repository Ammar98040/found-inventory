from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.db import models as db_models
from .models import Product, Location, Warehouse, AuditLog, DailyReportArchive, Order, UserProfile, UserActivityLog
from .decorators import admin_required, staff_required, exclude_maintenance, exclude_admin_dashboard, get_user_type, is_admin
import json
from django.core import serializers
from datetime import datetime, timedelta
from django.utils import timezone


@login_required
def home(request):
    """الصفحة الرئيسية - البحث عن المنتجات"""
    user = request.user
    user_profile = None
    user_type_display = 'موظف'
    
    if hasattr(user, 'user_profile'):
        user_profile = user.user_profile
        user_type_display = user_profile.get_user_type_display()
    elif user.is_superuser:
        user_type_display = 'مسؤول'
    
    return render(request, 'inventory_app/home.html', {
        'user': user,
        'user_profile': user_profile,
        'user_type_display': user_type_display
    })


@csrf_exempt
@transaction.atomic
def confirm_products(request):
    """تأكيد أخذ المنتجات وخصم الكميات"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            products_list = data.get('products', [])
            recipient_name = data.get('recipient_name', '').strip()
            
            # Extract product numbers for efficient batch query
            product_numbers = [item.get('number', '').strip() for item in products_list if item.get('number', '').strip()]
            
            # Optimize: Batch query with select_for_update to lock rows
            products_dict = {
                p.product_number: p 
                for p in Product.objects.filter(product_number__in=product_numbers).select_for_update()
            }
            
            updated_products = []
            
            for item in products_list:
                product_number = item.get('number', '').strip()
                requested_quantity = int(item.get('quantity', 0))
                
                # Fast lookup
                product = products_dict.get(product_number)
                
                if not product:
                    return JsonResponse({
                        'success': False,
                        'error': f'المنتج {product_number} غير موجود'
                    })
                
                old_quantity = product.quantity
                
                # إذا كانت الكمية المطلوبة 0، نسجل العملية فقط بدون خصم
                if requested_quantity == 0:
                    # تسجيل العملية في السجل بدون خصم الكمية
                    AuditLog.objects.create(
                        action='quantity_taken',
                        product=product,
                        product_number=product_number,
                        quantity_before=old_quantity,
                        quantity_after=old_quantity,
                        quantity_change=0,
                        notes=f'تم البحث عن المنتج بدون سحب كمية (كمية 0)',
                        user=request.user.username if request.user.is_authenticated else 'Guest'
                    )
                    
                    updated_products.append({
                        'product_number': product_number,
                        'old_quantity': old_quantity,
                        'new_quantity': old_quantity,
                        'quantity_taken': 0
                    })
                # إذا كانت الكمية المطلوبة أكبر من 0، نتحقق ونخصم
                elif product.quantity >= requested_quantity:
                    product.quantity -= requested_quantity
                    product.save()
                    
                    # تسجيل العملية في السجل
                    AuditLog.objects.create(
                        action='quantity_taken',
                        product=product,
                        product_number=product_number,
                        quantity_before=old_quantity,
                        quantity_after=product.quantity,
                        quantity_change=-requested_quantity,
                        notes=f'تم سحب {requested_quantity} من المنتج',
                        user=request.user.username if request.user.is_authenticated else 'Guest'
                    )
                    
                    updated_products.append({
                        'product_number': product_number,
                        'old_quantity': old_quantity,
                        'new_quantity': product.quantity,
                        'quantity_taken': requested_quantity
                    })
                else:
                    return JsonResponse({
                        'success': False,
                        'error': f'الكمية غير كافية للمنتج {product_number}'
                    })
            
            # حفظ الطلبية في السجل
            if updated_products:
                from datetime import datetime
                import random
                import string
                
                # إنشاء رقم طلبية فريد
                order_number = f"ORD-{datetime.now().strftime('%Y%m%d%H%M%S')}-{''.join(random.choices(string.ascii_uppercase + string.digits, k=4))}"
                
                # حساب الإجماليات
                total_products = len(updated_products)
                total_quantities = sum(p['quantity_taken'] for p in updated_products)
                
                # حفظ الطلبية
                order = Order.objects.create(
                    order_number=order_number,
                    products_data=updated_products,
                    total_products=total_products,
                    total_quantities=total_quantities,
                    recipient_name=recipient_name or None,
                    user=request.user.username if request.user.is_authenticated else 'Guest'
                )
                
                return JsonResponse({
                    'success': True,
                    'updated_products': updated_products,
                    'message': f'تم خصم {len(updated_products)} منتج',
                    'order_number': order_number
                })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)


@csrf_exempt
def search_products(request):
    """البحث عن المنتجات من خلال أرقامهم"""
    if request.method == 'POST':
        data = json.loads(request.body)
        products_list = data.get('products', [])
        
        # Extract product numbers for efficient batch query
        product_numbers = [item.get('product_number', '').strip() for item in products_list if item.get('product_number', '').strip()]
        
        # Optimize: Use select_related to avoid N+1 queries and batch query
        products = Product.objects.filter(product_number__in=product_numbers).select_related('location')
        
        # Create a dictionary for fast lookup
        products_dict = {p.product_number: p for p in products}
        
        results = []
        
        for item in products_list:
            product_number = item.get('product_number', '').strip()
            requested_quantity = int(item.get('quantity', 0))
            
            if not product_number:
                continue
            
            # Fast lookup from dictionary
            product = products_dict.get(product_number)
            
            if product:
                locations_data = []
                if product.location:
                    grid_pos = product.location.get_grid_position()
                    locations_data.append({
                        'id': product.location.id,
                        'full_location': product.location.full_location,
                        'row': product.location.row,
                        'column': product.location.column,
                        'x': grid_pos['x'],
                        'y': grid_pos['y'],
                        'notes': product.location.notes,
                    })
                
                result = {
                    'product_number': product.product_number,
                    'name': product.name,
                    'category': product.category,
                    'quantity': product.quantity,
                    'locations': locations_data,
                    'found': True,
                }
                
                if requested_quantity > 0:
                    result['requested_quantity'] = requested_quantity
                
                results.append(result)
            else:
                results.append({
                    'product_number': product_number,
                    'requested_quantity': requested_quantity,
                    'found': False,
                    'error': 'المنتج غير موجود في قاعدة البيانات'
                })
        
        return JsonResponse({'results': results}, json_dumps_params={'ensure_ascii': False})
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)


@require_http_methods(["GET"])
def get_products_list(request):
    products = Product.objects.all()[:100]
    products_data = [{'number': p.product_number, 'name': p.name} for p in products]
    return JsonResponse({'products': products_data}, json_dumps_params={'ensure_ascii': False})


def manage_warehouse(request):
    """صفحة إدارة المستودع"""
    warehouse = Warehouse.objects.first()
    if not warehouse:
        warehouse = Warehouse.objects.create(name='المستودع الرئيسي', rows_count=6, columns_count=15)
    
    return render(request, 'inventory_app/manage_warehouse.html', {'warehouse': warehouse})


@require_http_methods(["GET"])
def get_warehouse_grid(request):
    """الحصول على شبكة المستودع"""
    warehouse = Warehouse.objects.first()
    if not warehouse:
        return JsonResponse({'error': 'لا يوجد مستودع'}, status=404)
    
    # Optimize query with prefetch_related
    locations = Location.objects.filter(warehouse=warehouse).prefetch_related('products')
    grid_data = {}
    
    for location in locations:
        key = f"{location.row},{location.column}"
        # Use prefetched products
        products = list(location.products.all())
        grid_data[key] = {
            'row': location.row,
            'column': location.column,
            'notes': location.notes,
            'is_active': location.is_active,
            'has_products': len(products) > 0,
            'products': [p.product_number for p in products],
        }
    
    return JsonResponse({
        'rows': warehouse.rows_count,
        'columns': warehouse.columns_count,
        'grid': grid_data
    })


@require_http_methods(["POST"])
@csrf_exempt
def add_row(request):
    """إضافة صف/صفوف جديدة"""
    warehouse = Warehouse.objects.first()
    if not warehouse:
        return JsonResponse({'error': 'لا يوجد مستودع'}, status=404)
    
    try:
        data = json.loads(request.body)
        count = int(data.get('count', 1))  # عدد الصفوف المراد إضافتها
        
        if count < 1 or count > 50:
            return JsonResponse({'error': 'العدد يجب أن يكون بين 1 و 50'}, status=400)
        
        with transaction.atomic():
            rows_added = 0
            for _ in range(count):
                warehouse.rows_count += 1
                warehouse.save()
                
                for col in range(1, warehouse.columns_count + 1):
                    Location.objects.create(
                        warehouse=warehouse,
                        row=warehouse.rows_count,
                        column=col
                    )
                rows_added += 1
            
            return JsonResponse({
                'success': True,
                'new_rows_count': warehouse.rows_count,
                'rows_added': rows_added
            })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@require_http_methods(["POST"])
@csrf_exempt
def add_column(request):
    """إضافة عمود/أعمدة جديدة"""
    warehouse = Warehouse.objects.first()
    if not warehouse:
        return JsonResponse({'error': 'لا يوجد مستودع'}, status=404)
    
    try:
        data = json.loads(request.body)
        count = int(data.get('count', 1))  # عدد الأعمدة المراد إضافتها
        
        if count < 1 or count > 50:
            return JsonResponse({'error': 'العدد يجب أن يكون بين 1 و 50'}, status=400)
        
        with transaction.atomic():
            columns_added = 0
            for _ in range(count):
                warehouse.columns_count += 1
                warehouse.save()
                
                for row in range(1, warehouse.rows_count + 1):
                    Location.objects.create(
                        warehouse=warehouse,
                        row=row,
                        column=warehouse.columns_count
                    )
                columns_added += 1
            
            return JsonResponse({
                'success': True,
                'new_columns_count': warehouse.columns_count,
                'columns_added': columns_added
            })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@require_http_methods(["POST"])
@csrf_exempt
def delete_row(request):
    """حذف صف/صفوف"""
    warehouse = Warehouse.objects.first()
    if not warehouse:
        return JsonResponse({'error': 'لا يوجد مستودع'}, status=404)
    
    try:
        data = json.loads(request.body)
        count = int(data.get('count', 1))  # عدد الصفوف المراد حذفها
        
        if count < 1 or count > warehouse.rows_count:
            return JsonResponse({'error': f'العدد يجب أن يكون بين 1 و {warehouse.rows_count}'}, status=400)
        
        with transaction.atomic():
            rows_deleted = 0
            for _ in range(count):
                if warehouse.rows_count > 0:
                    # حذف آخر صف
                    Location.objects.filter(warehouse=warehouse, row=warehouse.rows_count).delete()
                    warehouse.rows_count -= 1
                    warehouse.save()
                    rows_deleted += 1
                else:
                    break
            
            return JsonResponse({
                'success': True,
                'new_rows_count': warehouse.rows_count,
                'rows_deleted': rows_deleted
            })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@require_http_methods(["POST"])
@csrf_exempt
def delete_column(request):
    """حذف عمود/أعمدة"""
    warehouse = Warehouse.objects.first()
    if not warehouse:
        return JsonResponse({'error': 'لا يوجد مستودع'}, status=404)
    
    try:
        data = json.loads(request.body)
        count = int(data.get('count', 1))  # عدد الأعمدة المراد حذفها
        
        if count < 1 or count > warehouse.columns_count:
            return JsonResponse({'error': f'العدد يجب أن يكون بين 1 و {warehouse.columns_count}'}, status=400)
        
        with transaction.atomic():
            columns_deleted = 0
            for _ in range(count):
                if warehouse.columns_count > 0:
                    # حذف آخر عمود
                    Location.objects.filter(warehouse=warehouse, column=warehouse.columns_count).delete()
                    warehouse.columns_count -= 1
                    warehouse.save()
                    columns_deleted += 1
                else:
                    break
            
            return JsonResponse({
                'success': True,
                'new_columns_count': warehouse.columns_count,
                'columns_deleted': columns_deleted
            })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


def warehouse_dashboard(request):
    """لوحة تحكم المستودع"""
    # Use select_related for related data
    warehouse = Warehouse.objects.first()
    products_count = Product.objects.count()
    locations_count = Location.objects.count()
    total_capacity = warehouse.rows_count * warehouse.columns_count if warehouse else 0
    
    return render(request, 'inventory_app/dashboard.html', {
        'warehouse': warehouse,
        'products_count': products_count,
        'locations_count': locations_count,
        'total_capacity': total_capacity
    })


def products_list(request):
    """قائمة جميع المنتجات"""
    # Optimize query with select_related to avoid N+1 queries
    products = Product.objects.select_related('location').all().order_by('product_number')
    
    # احسب العدد الإجمالي قبل أي فلترة
    total_count = products.count()
    
    search = request.GET.get('search', '')
    if search:
        products = products.filter(
            product_number__icontains=search
        ) | products.filter(
            name__icontains=search
        )
    
    # احسب العدد بعد الفلترة (إذا كان هناك بحث)
    filtered_count = products.count() if search else total_count
    
    # عرض جميع المنتجات بدون تقسيم
    return render(request, 'inventory_app/products_list.html', {
        'products': products,
        'search': search,
        'total_count': total_count,
        'filtered_count': filtered_count
    })


def product_detail(request, product_id):
    """تفاصيل منتج"""
    product = get_object_or_404(Product, id=product_id)
    return render(request, 'inventory_app/product_detail.html', {'product': product})


def product_add(request):
    """إضافة منتج جديد"""
    if request.method == 'POST':
        try:
            product = Product.objects.create(
                product_number=request.POST.get('product_number'),
                name=request.POST.get('name', ''),
                category=request.POST.get('category', ''),
                description=request.POST.get('description', ''),
                quantity=int(request.POST.get('quantity', 0) or 0)
            )
            
            # تسجيل العملية في السجل
            AuditLog.objects.create(
                action='added',
                product=product,
                product_number=product.product_number,
                quantity_before=0,
                quantity_after=product.quantity,
                quantity_change=product.quantity,
                notes=f'تم إضافة منتج جديد: {product.name}',
                user=request.user.username if request.user.is_authenticated else 'Guest'
            )
            
            messages.success(request, 'تم إضافة المنتج بنجاح')
            return redirect('inventory_app:product_detail', product_id=product.id)
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    return render(request, 'inventory_app/product_add.html')


def product_edit(request, product_id):
    """تعديل منتج"""
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        try:
            # حفظ القيم القديمة
            old_product_number = product.product_number
            old_name = product.name
            old_category = product.category
            old_quantity = product.quantity
            
            # التعديل
            new_product_number = request.POST.get('product_number')
            new_name = request.POST.get('name', '')
            new_category = request.POST.get('category', '')
            new_quantity = int(request.POST.get('quantity', 0) or 0)
            
            product.product_number = new_product_number
            product.name = new_name
            product.category = new_category
            product.description = request.POST.get('description', '')
            product.quantity = new_quantity
            product.save()
            
            # تسجيل التغييرات في السجل
            changes = []
            if old_product_number != new_product_number:
                changes.append(f'رقم المنتج: {old_product_number} → {new_product_number}')
            if old_name != new_name:
                changes.append(f'الاسم: {old_name} → {new_name}')
            if old_category != new_category:
                changes.append(f'الفئة: {old_category} → {new_category}')
            if old_quantity != new_quantity:
                changes.append(f'الكمية: {old_quantity} → {new_quantity}')
            
            if changes:
                AuditLog.objects.create(
                    action='updated',
                    product=product,
                    product_number=product.product_number,
                    quantity_before=old_quantity,
                    quantity_after=new_quantity,
                    quantity_change=new_quantity - old_quantity,
                    notes='تغييرات: ' + ' | '.join(changes),
                    user=request.user.username if request.user.is_authenticated else 'Guest'
                )
            
            messages.success(request, 'تم تعديل المنتج بنجاح')
            return redirect('inventory_app:product_detail', product_id=product.id)
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    return render(request, 'inventory_app/product_edit.html', {'product': product})


@csrf_exempt
def product_delete(request, product_id):
    """حذف منتج"""
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        try:
            product_number = product.product_number
            quantity = product.quantity
            name = product.name
            
            # تسجيل العملية قبل الحذف
            AuditLog.objects.create(
                action='deleted',
                product=product,
                product_number=product_number,
                quantity_before=quantity,
                quantity_after=0,
                quantity_change=-quantity,
                notes=f'تم حذف المنتج: {name}',
                user=request.user.username if request.user.is_authenticated else 'Guest'
            )
            
            product.delete()
            
            # إرجاع JSON للطلبات AJAX
            if request.content_type == 'application/json':
                return JsonResponse({'success': True, 'message': f'تم حذف المنتج {product_number}'}, json_dumps_params={'ensure_ascii': False})
            
            messages.success(request, f'تم حذف المنتج {product_number}')
            return redirect('inventory_app:products_list')
        except Exception as e:
            if request.content_type == 'application/json':
                return JsonResponse({'success': False, 'error': str(e)}, json_dumps_params={'ensure_ascii': False})
            messages.error(request, f'خطأ في حذف المنتج: {str(e)}')
            return redirect('inventory_app:product_detail', product_id=product.id)
    
    return render(request, 'inventory_app/product_delete.html', {'product': product})


@csrf_exempt
def delete_products_bulk(request):
    """حذف منتجات متعددة دفعة واحدة"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            product_ids = data.get('product_ids', [])
            
            if not product_ids:
                return JsonResponse({
                    'success': False,
                    'error': 'لم يتم تحديد أي منتجات'
                }, json_dumps_params={'ensure_ascii': False})
            
            deleted_count = 0
            deleted_products = []
            
            with transaction.atomic():
                products = Product.objects.filter(id__in=product_ids)
                
                for product in products:
                    product_number = product.product_number
                    quantity = product.quantity
                    name = product.name
                    
                    # تسجيل العملية قبل الحذف
                    AuditLog.objects.create(
                        action='deleted',
                        product=product,
                        product_number=product_number,
                        quantity_before=quantity,
                        quantity_after=0,
                        quantity_change=-quantity,
                        notes=f'حذف جماعي: {name}',
                        user=request.user.username if request.user.is_authenticated else 'Guest'
                    )
                    
                    deleted_products.append(product_number)
                
                # حذف المنتجات
                products.delete()
                deleted_count = len(deleted_products)
            
            return JsonResponse({
                'success': True,
                'message': f'تم حذف {deleted_count} منتج بنجاح',
                'deleted_count': deleted_count,
                'deleted_products': deleted_products[:10]  # أول 10 فقط
            }, json_dumps_params={'ensure_ascii': False})
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, json_dumps_params={'ensure_ascii': False})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)


@csrf_exempt
def move_product_with_shift(request, product_id):
    """نقل منتج لموقع معين وإعادة ترتيب باقي المنتجات في نفس العمود"""
    try:
        product = get_object_or_404(Product, id=product_id)
        data = json.loads(request.body)
        
        new_location_str = data.get('new_location', '')
        if not new_location_str:
            return JsonResponse({'success': False, 'error': 'الموقع الجديد مطلوب'}, json_dumps_params={'ensure_ascii': False})
        
        # تحليل الموقع الجديد (مثال: R15C4)
        import re
        match = re.match(r'R(\d+)C(\d+)', new_location_str)
        if not match:
            return JsonResponse({'success': False, 'error': 'تنسيق الموقع غير صحيح'}, json_dumps_params={'ensure_ascii': False})
        
        new_row = int(match.group(1))
        new_column = int(match.group(2))
        
        warehouse = Warehouse.objects.first()
        if not warehouse:
            return JsonResponse({'success': False, 'error': 'لا يوجد مستودع'}, json_dumps_params={'ensure_ascii': False})
        
        # التحقق من أن العمود الجديد لا يحتوي على عدد صفوف كافي
        if new_row > warehouse.rows_count:
            # إضافة صفوف إضافية
            for row in range(warehouse.rows_count + 1, new_row + 1):
                warehouse.rows_count += 1
                warehouse.save()
                
                # إنشاء المواقع للعمود الجديد
                for col in range(1, warehouse.columns_count + 1):
                    Location.objects.get_or_create(
                        warehouse=warehouse,
                        row=row,
                        column=col,
                        defaults={'is_active': True}
                    )
        
        with transaction.atomic():
            old_location = product.location
            old_row = old_location.row if old_location else None
            old_column = old_location.column if old_location else None
            
            # نقل المنتج للموقع الجديد أولاً
            new_location = Location.objects.get(warehouse=warehouse, row=new_row, column=new_column)
            product.location = new_location
            product.save()
            
            # تسجيل العملية للمنتج المنقول
            old_location_str = old_location.full_location if old_location else 'بدون موقع'
            AuditLog.objects.create(
                action='location_assigned',
                product=product,
                product_number=product.product_number,
                quantity_before=product.quantity,
                quantity_after=product.quantity,
                quantity_change=0,
                notes=f'نقل مع إعادة ترتيب: {old_location_str} → {new_location_str}',
                user=request.user.username if request.user.is_authenticated else 'Guest'
            )
            
            # إذا كان النقل في نفس العمود، نحتاج لإعادة ترتيب المنتجات
            if old_location and old_column == new_column:
                # سيناريو 1: نقل المنتج لأسفل في نفس العمود (من R10 إلى R15)
                if old_row < new_row:
                    # جلب المنتجات التي كانت بين الموقع القديم والجديد
                    products_to_shift_up = Product.objects.filter(
                        location__warehouse=warehouse,
                        location__column=old_column,
                        location__row__gt=old_row,
                        location__row__lte=new_row
                    ).exclude(id=product.id).select_related('location').order_by('location__row')
                    
                    # نقل هذه المنتجات صف واحد لأعلى
                    for prod in products_to_shift_up:
                        old_loc = prod.location
                        # نقل المنتج لصف واحد أقل
                        if old_loc.row > 1:
                            new_loc = Location.objects.get(
                                warehouse=warehouse,
                                row=old_loc.row - 1,
                                column=old_loc.column
                            )
                            prod.location = new_loc
                            prod.save()
                            
                            AuditLog.objects.create(
                                action='location_assigned',
                                product=prod,
                                product_number=prod.product_number,
                                quantity_before=prod.quantity,
                                quantity_after=prod.quantity,
                                quantity_change=0,
                                notes=f'إعادة ترتيب تلقائي: {old_loc.full_location} → {new_loc.full_location}',
                                user=request.user.username if request.user.is_authenticated else 'Guest'
                            )
                
                # سيناريو 2: نقل المنتج لأعلى في نفس العمود (من R15 إلى R10)
                elif old_row > new_row:
                    # جلب المنتجات التي كانت بين الموقع الجديد والقديم
                    products_to_shift_down = Product.objects.filter(
                        location__warehouse=warehouse,
                        location__column=old_column,
                        location__row__gte=new_row,
                        location__row__lt=old_row
                    ).exclude(id=product.id).select_related('location').order_by('location__row')
                    
                    # نقل هذه المنتجات صف واحد لأسفل
                    for prod in products_to_shift_down:
                        old_loc = prod.location
                        # نقل المنتج لصف واحد أكثر
                        new_row_for_prod = old_loc.row + 1
                        
                        # التحقق من وجود صفوف كافية
                        if new_row_for_prod > warehouse.rows_count:
                            warehouse.rows_count += 1
                            warehouse.save()
                            
                            # إنشاء المواقع للصف الجديد
                            for col in range(1, warehouse.columns_count + 1):
                                Location.objects.get_or_create(
                                    warehouse=warehouse,
                                    row=warehouse.rows_count,
                                    column=col,
                                    defaults={'is_active': True}
                                )
                        
                        new_loc = Location.objects.get(
                            warehouse=warehouse,
                            row=new_row_for_prod,
                            column=old_loc.column
                        )
                        prod.location = new_loc
                        prod.save()
                        
                        AuditLog.objects.create(
                            action='location_assigned',
                            product=prod,
                            product_number=prod.product_number,
                            quantity_before=prod.quantity,
                            quantity_after=prod.quantity,
                            quantity_change=0,
                            notes=f'إعادة ترتيب تلقائي: {old_loc.full_location} → {new_loc.full_location}',
                            user=request.user.username if request.user.is_authenticated else 'Guest'
                        )
            
            # إذا كان النقل لعمود مختلف، لا حاجة لإعادة ترتيب
            shifted_count = 0 if not old_location or old_column != new_column else 1
            
            return JsonResponse({
                'success': True,
                'message': f'تم نقل المنتج {product.product_number} بنجاح من {old_location_str} إلى {new_location_str}'
            }, json_dumps_params={'ensure_ascii': False})
            
    except Location.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'الموقع المطلوب غير موجود'}, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, json_dumps_params={'ensure_ascii': False})


def assign_location_to_product(request, product_id):
    """ربط منتج بموقع واحد فقط"""
    # Optimize with select_related
    product = get_object_or_404(Product.objects.select_related('location'), id=product_id)
    warehouse = Warehouse.objects.first()
    
    # إنشاء جميع المواقع المفقودة
    if warehouse:
        locations_created = 0
        with transaction.atomic():
            for row in range(1, warehouse.rows_count + 1):
                for col in range(1, warehouse.columns_count + 1):
                    location, created = Location.objects.get_or_create(
                        warehouse=warehouse,
                        row=row,
                        column=col,
                        defaults={
                            'is_active': True
                        }
                    )
                    if created:
                        locations_created += 1
    
    if request.method == 'POST':
        location_id = request.POST.get('location')
        if location_id:
            new_location = Location.objects.get(id=location_id)
            old_location = product.location
            
            # التحقق من أن الموقع غير مشغول بمنتج آخر - use exists() for efficiency
            if new_location.products.exclude(id=product.id).exists():
                messages.error(request, 'هذا الموقع مشغول بمنتج آخر! اختر موقعاً آخر')
            else:
                product.location = new_location
                product.save()
                
                # تسجيل العملية
                old_location_str = old_location.full_location if old_location else 'بدون موقع'
                new_location_str = new_location.full_location
                
                AuditLog.objects.create(
                    action='location_assigned',
                    product=product,
                    product_number=product.product_number,
                    quantity_before=product.quantity,
                    quantity_after=product.quantity,
                    quantity_change=0,
                    notes=f'تغيير الموقع: {old_location_str} → {new_location_str}',
                    user=request.user.username if request.user.is_authenticated else 'Guest'
                )
                
                messages.success(request, f'تم ربط المنتج بالموقع {new_location_str} بنجاح')
        else:
            # إلغاء الربط
            if product.location:
                old_location_str = product.location.full_location
                product.location = None
                product.save()
                
                AuditLog.objects.create(
                    action='location_removed',
                    product=product,
                    product_number=product.product_number,
                    quantity_before=product.quantity,
                    quantity_after=product.quantity,
                    quantity_change=0,
                    notes=f'إلغاء ربط الموقع: {old_location_str}',
                    user=request.user.username if request.user.is_authenticated else 'Guest'
                )
                
                messages.success(request, 'تم إلغاء ربط الموقع')
        
        return redirect('inventory_app:product_detail', product_id=product.id)
    
    # Optimize query
    all_locations = Location.objects.filter(warehouse=warehouse).select_related('warehouse').order_by('row', 'column')
    
    # تحديد الأماكن الشاغرة والفارغة - optimize with values_list
    occupied_locations = set(
        Location.objects.filter(products__isnull=False)
        .values_list('id', flat=True)
    )
    
    return render(request, 'inventory_app/assign_location.html', {
        'product': product,
        'all_locations': all_locations,
        'current_location': product.location,
        'warehouse': warehouse,
        'occupied_locations': occupied_locations
    })


def warehouses_list(request):
    """قائمة المستودعات"""
    warehouses = Warehouse.objects.all()
    return render(request, 'inventory_app/warehouses_list.html', {'warehouses': warehouses})


def warehouse_detail(request, warehouse_id):
    """تفاصيل مستودع"""
    warehouse = get_object_or_404(Warehouse, id=warehouse_id)
    locations_count = Location.objects.filter(warehouse=warehouse).count()
    return render(request, 'inventory_app/warehouse_detail.html', {
        'warehouse': warehouse,
        'locations_count': locations_count
    })


def locations_list(request):
    """قائمة جميع الأماكن"""
    warehouse = Warehouse.objects.first()
    
    if warehouse:
        # إنشاء جميع المواقع المفقودة
        locations_created = 0
        with transaction.atomic():
            for row in range(1, warehouse.rows_count + 1):
                for col in range(1, warehouse.columns_count + 1):
                    location, created = Location.objects.get_or_create(
                        warehouse=warehouse,
                        row=row,
                        column=col,
                        defaults={
                            'is_active': True
                        }
                    )
                    if created:
                        locations_created += 1
        
        if locations_created > 0:
            print(f'Created {locations_created} missing locations')
    
    # Optimize with prefetch_related and select_related
    # الحصول على جميع المواقع للشبكة (بدون pagination)
    all_locations = Location.objects.filter(warehouse=warehouse).select_related('warehouse').prefetch_related('products').order_by('row', 'column')
    
    # للحصول على paginated locations للقائمة (إن وجدت في المستقبل)
    search = request.GET.get('search', '')
    if search:
        all_locations = all_locations.filter(notes__icontains=search)
    
    # Add pagination
    from django.core.paginator import Paginator
    paginator = Paginator(all_locations, 100)  # Show 100 locations per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'inventory_app/locations_list.html', {
        'locations': all_locations,  # تمرير جميع المواقع للشبكة
        'page_obj': page_obj,
        'warehouse': warehouse,
        'search': search
    })


@require_http_methods(["GET"])
def get_stats(request):
    """API للحصول على إحصائيات النظام"""
    try:
        from .models import Product, Location, Warehouse
        
        # إحصائيات المنتجات
        products_count = Product.objects.count()
        # المنتجات التي لها موقع (location ليست null)
        products_with_locations = Product.objects.filter(location__isnull=False).count()
        products_without_locations = products_count - products_with_locations
        
        # إحصائيات الأماكن
        locations_count = Location.objects.count()
        warehouse = Warehouse.objects.first()
        
        if warehouse:
            total_capacity = warehouse.rows_count * warehouse.columns_count
            # المواقع المشغولة (التي تحتوي على منتجات)
            # استخدام values_list للحصول على الـ IDs المميزة للأماكن المشغولة
            occupied_locations_ids = Product.objects.filter(location__isnull=False).values_list('location_id', flat=True).distinct()
            occupied_locations = len(occupied_locations_ids) if occupied_locations_ids else 0
            empty_locations = total_capacity - occupied_locations
        else:
            total_capacity = 0
            occupied_locations = 0
            empty_locations = 0
        
        return JsonResponse({
            # المنتجات
            'products_count': products_count,
            'products_with_locations': products_with_locations,
            'products_without_locations': products_without_locations,
            # الأماكن
            'locations_count': locations_count,
            'total_capacity': total_capacity,
            'occupied_locations': occupied_locations,
            'empty_locations': empty_locations,
            # معلومات المستودع
            'warehouse_rows': warehouse.rows_count if warehouse else 0,
            'warehouse_columns': warehouse.columns_count if warehouse else 0,
        }, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def quick_search_products(request):
    """API للبحث السريع في المنتجات"""
    try:
        from .models import Product
        query = request.GET.get('q', '').strip()
        
        if not query:
            return JsonResponse([], safe=False, json_dumps_params={'ensure_ascii': False})
        
        # البحث في رقم المنتج واسمه
        products = Product.objects.filter(
            product_number__icontains=query
        ) | Product.objects.filter(
            name__icontains=query
        )
        
        products = products[:10]  # أول 10 نتائج
        
        results = []
        for product in products:
            results.append({
                'id': product.id,
                'product_number': product.product_number,
                'name': product.name,
                'location': product.location.full_location if product.location else None,
                'quantity': product.quantity
            })
        
        return JsonResponse(results, safe=False, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def quick_search_locations(request):
    """API للبحث السريع في الأماكن"""
    try:
        from .models import Location
        query = request.GET.get('q', '').strip()
        
        if not query:
            return JsonResponse([], safe=False, json_dumps_params={'ensure_ascii': False})
        
        # البحث في المواقع (R1C1, R2C3, etc.)
        locations = Location.objects.filter(
            row__icontains=query
        ) | Location.objects.filter(
            column__icontains=query
        )
        
        locations = locations[:10]  # أول 10 نتائج
        
        results = []
        for location in locations:
            has_product = location.products.exists()
            results.append({
                'id': location.id,
                'full_location': location.full_location,
                'row': location.row,
                'column': location.column,
                'has_product': has_product,
                'warehouse': location.warehouse.name if location.warehouse else ''
            })
        
        return JsonResponse(results, safe=False, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def audit_logs(request):
    """صفحة عرض سجلات العمليات"""
    from django.core.paginator import Paginator
    from django.db.models import Count

    # الاستعلام الأساسي + تحسينات
    base_qs = AuditLog.objects.select_related('product').all()

    search = request.GET.get('search', '')
    action_filter = request.GET.get('action', '')

    if search:
        base_qs = base_qs.filter(product_number__icontains=search)

    if action_filter:
        base_qs = base_qs.filter(action=action_filter)

    # إحصائيات كاملة قبل التقسيم لصفحات
    total_count = base_qs.count()
    counts_qs = base_qs.values('action').annotate(count=Count('id'))
    action_counts = {row['action']: row['count'] for row in counts_qs}

    # ترتيب لضمان تجميع منطقي داخل الصفحة (ثم الأحدث داخل كل نوع)
    ordered_qs = base_qs.order_by('action', '-created_at')

    # التقسيم لصفحات
    paginator = Paginator(ordered_qs, 100)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'inventory_app/audit_logs.html', {
        'logs': page_obj,
        'page_obj': page_obj,
        'search': search,
        'action_filter': action_filter,
        'total_count': total_count,
        'action_counts': action_counts,
    })


# ========== تصدير البيانات ==========

def export_products_pdf(request):
    """تصدير قائمة المنتجات إلى PDF احترافي مع دعم كامل للعربية باستخدام Playwright"""
    from django.http import HttpResponse
    from playwright.sync_api import sync_playwright
    from datetime import datetime
    import io
    
    try:
        # جلب المنتجات
        products = Product.objects.select_related('location').all().order_by('product_number')
        
        # تاريخ التقرير
        now = datetime.now()
        
        # إنشاء HTML محتوى
        html_content = f'''
<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
    <meta charset="UTF-8">
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', 'Arial', 'Tahoma', sans-serif;
            font-size: 10pt;
            direction: rtl;
            padding: 20px;
        }}
        
        .header {{
            text-align: center;
            color: #667eea;
            font-size: 28pt;
            font-weight: bold;
            margin-bottom: 20px;
        }}
        
        .info {{
            text-align: center;
            margin-bottom: 20px;
            font-size: 11pt;
            color: #374151;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 0 auto;
        }}
        
        th {{
            background-color: #667eea;
            color: white;
            padding: 12px 8px;
            text-align: right;
            font-weight: bold;
            border: 1px solid #555;
            font-size: 11pt;
        }}
        
        td {{
            padding: 8px;
            border: 1px solid #ddd;
            text-align: right;
            font-size: 9pt;
        }}
        
        tr:nth-child(even) {{
            background-color: #f8fafc;
        }}
        
        .summary {{
            margin-top: 20px;
            text-align: right;
            font-weight: bold;
            font-size: 12pt;
            color: #374151;
        }}
    </style>
</head>
<body>
    <div class="header">قائمة المنتجات</div>
    
    <div class="info">
        <strong>التاريخ:</strong> {now.strftime("%Y-%m-%d")} | 
        <strong>الوقت:</strong> {now.strftime("%H:%M")}
    </div>
    
    <table>
        <thead>
            <tr>
                <th>#</th>
                <th>رقم المنتج</th>
                <th>الاسم</th>
                <th>الفئة</th>
                <th>الكمية</th>
                <th>الموقع</th>
            </tr>
        </thead>
        <tbody>
'''
        
        # إضافة صفوف المنتجات
        for idx, product in enumerate(products, start=1):
            location = product.location.full_location if product.location else 'بدون موقع'
            category = product.category if product.category else '-'
            
            html_content += f'''
            <tr>
                <td>{idx}</td>
                <td>{product.product_number}</td>
                <td>{product.name}</td>
                <td>{category}</td>
                <td>{product.quantity}</td>
                <td>{location}</td>
            </tr>
            '''
        
        # إغلاق HTML
        html_content += f'''
        </tbody>
    </table>
    
    <div class="summary">
        إجمالي المنتجات: {products.count()}
    </div>
</body>
</html>
'''
        
        # إنشاء PDF باستخدام Playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_content(html_content)
            
            pdf_bytes = page.pdf(
                format='A4',
                landscape=True,
                margin={'top': '1cm', 'right': '1cm', 'bottom': '1cm', 'left': '1cm'}
            )
            
            browser.close()
        
        # إرجاع الاستجابة
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="products_list.pdf"'
        
        return response
        
    except Exception as e:
        import traceback
        error_msg = f'خطأ في إنشاء PDF: {str(e)}\n{traceback.format_exc()}'
        return HttpResponse(error_msg, content_type='text/plain')


def export_products_excel(request):
    """تصدير قائمة المنتجات إلى Excel"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # إنشاء workbook جديد
    wb = Workbook()
    ws = wb.active
    ws.title = "المنتجات"
    
    # تنسيق النمط العربي
    arabic_font = Font(name='Arial', size=12, bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # رأس الجدول
    headers = ['#', 'رقم المنتج', 'الاسم', 'الفئة', 'الكمية', 'الموقع', 'تاريخ الإضافة']
    ws.append(headers)
    
    # تنسيق رأس الجدول
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        cell.border = border
    
    # جلب المنتجات
    products = Product.objects.all().order_by('product_number')
    
    # إضافة البيانات
    for idx, product in enumerate(products, start=1):
        location_text = product.location.full_location if product.location else 'لا يوجد موقع'
        
        row_data = [
            idx,
            product.product_number,
            product.name,
            product.category or '',
            product.quantity,
            location_text,
            product.created_at.strftime('%Y-%m-%d')
        ]
        
        ws.append(row_data)
        
        # تنسيق الصف
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=idx + 1, column=col)
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            cell.border = border
            cell.font = Font(name='Arial', size=11)
    
    # ضبط عرض الأعمدة
    column_widths = [5, 15, 25, 15, 10, 12, 15]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # إعداد الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="قائمة_المنتجات.xlsx"'
    
    wb.save(response)
    return response


def daily_reports(request):
    """صفحة التقارير اليومية السريعة"""
    from datetime import datetime, timedelta
    
    # تحديد نطاق التاريخ (اليوم)
    today = datetime.now().date()
    
    # فحص وحفظ تلقائي للتقارير السابقة غير المحفوظة
    _auto_save_daily_reports()
    
    # إحصائيات اليوم (استخدام created_at بدلاً من timestamp)
    today_logs = AuditLog.objects.filter(created_at__date=today)
    
    # إحصائيات المنتجات
    products_added = today_logs.filter(action='added').count()
    products_updated = today_logs.filter(action='updated').count()
    products_deleted = today_logs.filter(action='deleted').count()
    quantities_taken = today_logs.filter(action='quantity_taken').count()
    locations_assigned = today_logs.filter(action='location_assigned').count()
    
    # المنتجات الجديدة اليوم
    new_products = Product.objects.filter(created_at__date=today).order_by('-created_at')[:10]
    
    # آخر العمليات
    recent_logs = today_logs.order_by('-created_at')[:20]
    
    # إحصائيات الكميات
    quantity_changes = today_logs.filter(action__in=['added', 'quantity_taken']).aggregate(
        total_added=db_models.Sum('quantity_change', filter=db_models.Q(quantity_change__gt=0)),
        total_removed=db_models.Sum('quantity_change', filter=db_models.Q(quantity_change__lt=0))
    )
    
    context = {
        'today': today,
        'products_added': products_added,
        'products_updated': products_updated,
        'products_deleted': products_deleted,
        'quantities_taken': quantities_taken,
        'locations_assigned': locations_assigned,
        'new_products': new_products,
        'recent_logs': recent_logs,
        'total_added': quantity_changes['total_added'] or 0,
        'total_removed': abs(quantity_changes['total_removed'] or 0),
    }
    
    return render(request, 'inventory_app/daily_reports.html', context)


def _auto_save_daily_reports():
    """حفظ تلقائي للتقارير اليومية السابقة غير المحفوظة"""
    from datetime import datetime, timedelta
    
    try:
        # التحقق من آخر 7 أيام للحفظ التلقائي
        today = datetime.now().date()
        
        for day_offset in range(7):
            check_date = today - timedelta(days=day_offset)
            
            # تخطي اليوم الحالي
            if check_date == today:
                continue
            
            # التحقق من عدم وجود تقرير محفوظ لهذا التاريخ
            existing_report = DailyReportArchive.objects.filter(report_date=check_date).first()
            if existing_report:
                continue  # التقرير محفوظ مسبقاً
            
            # جلب بيانات هذا التاريخ
            date_logs = AuditLog.objects.filter(created_at__date=check_date)
            
            # إذا كانت هناك عمليات في هذا اليوم ولم يكن محفوظ، احفظه
            if date_logs.exists():
                products_added = date_logs.filter(action='added').count()
                products_updated = date_logs.filter(action='updated').count()
                products_deleted = date_logs.filter(action='deleted').count()
                quantities_taken = date_logs.filter(action='quantity_taken').count()
                locations_assigned = date_logs.filter(action='location_assigned').count()
                
                new_products = Product.objects.filter(created_at__date=check_date).order_by('-created_at')[:10]
                recent_logs = date_logs.order_by('-created_at')[:20]
                
                quantity_changes = date_logs.filter(action__in=['added', 'quantity_taken']).aggregate(
                    total_added=db_models.Sum('quantity_change', filter=db_models.Q(quantity_change__gt=0)),
                    total_removed=db_models.Sum('quantity_change', filter=db_models.Q(quantity_change__lt=0))
                )
                
                # إعداد البيانات للـ JSON
                report_data = {
                    'new_products': [
                        {
                            'product_number': p.product_number,
                            'name': p.name,
                            'category': p.category,
                            'quantity': p.quantity,
                            'location': p.location.full_location if p.location else None
                        } for p in new_products
                    ],
                    'recent_logs': [
                        {
                            'action': log.action,
                            'product_number': log.product_number,
                            'notes': log.notes,
                            'created_at': log.created_at.strftime('%Y-%m-%d %H:%M')
                        } for log in recent_logs
                    ]
                }
                
                # إنشاء التقرير المحفوظ
                hijri_date = convert_to_hijri(check_date)
                
                DailyReportArchive.objects.create(
                    report_date=check_date,
                    hijri_date=hijri_date,
                    products_added=products_added,
                    products_updated=products_updated,
                    products_deleted=products_deleted,
                    quantities_taken=quantities_taken,
                    locations_assigned=locations_assigned,
                    total_added=quantity_changes['total_added'] or 0,
                    total_removed=abs(quantity_changes['total_removed'] or 0),
                    report_data=report_data,
                    is_auto_saved=True  # علامة لحفظ تلقائي
                )
    except Exception as e:
        # في حالة حدوث خطأ، لا توقف العملية
        pass


def convert_to_hijri(gregorian_date):
    """تحويل التاريخ الميلادي إلى هجري باستخدام حساب أدق"""
    try:
        from datetime import datetime
        import math
        
        # التاريخ المرجعي: 16 يوليو 622 ميلادي = 1 محرم 1 هجري
        gregorian_start = datetime(622, 7, 16)
        hijri_start = 1  # سنة 1 هجري
        
        # حساب الفرق بالأيام
        date_obj = datetime(gregorian_date.year, gregorian_date.month, gregorian_date.day)
        days_diff = (date_obj - gregorian_start).days
        
        # حساب السنة الهجرية (السنة الهجرية = 354.37 يوم في المتوسط)
        # مع تعديل أدق
        hijri_year = 1 + int(days_diff / 354.367)
        
        # حساب الأيام المتبقية منذ بداية السنة
        days_from_start_of_year = days_diff % 354
        
        # أشهر السنة الهجرية مع عدد أيامها (تقريبي)
        hijri_months = [
            ('محرم', 30), ('صفر', 29), ('ربيع الأول', 30), ('ربيع الآخر', 29),
            ('جمادى الأولى', 30), ('جمادى الآخرة', 29), ('رجب', 30), ('شعبان', 29),
            ('رمضان', 30), ('شوال', 29), ('ذو القعدة', 30), ('ذو الحجة', 29)
        ]
        
        # حساب الشهر واليوم
        remaining_days = days_from_start_of_year
        hijri_month = 1
        hijri_day = 1
        
        for month_name, month_days in hijri_months:
            if remaining_days < month_days:
                hijri_day = remaining_days + 1
                break
            remaining_days -= month_days
            hijri_month += 1
        
        # ضمان أن الشهر ضمن النطاق الصحيح
        if hijri_month > 12:
            hijri_month = 12
            hijri_day = min(hijri_day, 29)
        
        month_name = hijri_months[hijri_month - 1][0]
        
        # حساب السنة الهجرية بدقة أكبر باستخدام سنة كبيسة
        # السنة الهجرية الكبيسة لها 355 يوماً (3 سنوات في كل 8)
        leap_days = int(hijri_year / 30) * 11  # كل 30 سنة = 11 يوم إضافي
        adjusted_days = days_diff - leap_days
        
        # إعادة حساب السنة
        final_year = int(adjusted_days / 354.367) + 1
        
        return f"{hijri_day} {month_name} {final_year} هـ"
    except Exception as e:
        # في حالة الخطأ، رجع التاريخ الميلادي
        return f"{gregorian_date.strftime('%Y-%m-%d')}"


def save_daily_report(request):
    """حفظ التقرير اليومي"""
    from datetime import datetime
    
    if request.method == 'POST':
        try:
            today = datetime.now().date()
            
            # التحقق من عدم وجود تقرير محفوظ لهذا اليوم
            existing_report = DailyReportArchive.objects.filter(report_date=today).first()
            if existing_report:
                return JsonResponse({
                    'success': False,
                    'message': 'تم حفظ تقرير اليوم مسبقاً'
                })
            
            # جلب بيانات اليوم
            today_logs = AuditLog.objects.filter(created_at__date=today)
            
            products_added = today_logs.filter(action='added').count()
            products_updated = today_logs.filter(action='updated').count()
            products_deleted = today_logs.filter(action='deleted').count()
            quantities_taken = today_logs.filter(action='quantity_taken').count()
            locations_assigned = today_logs.filter(action='location_assigned').count()
            
            new_products = Product.objects.filter(created_at__date=today).order_by('-created_at')[:10]
            recent_logs = today_logs.order_by('-created_at')[:20]
            
            quantity_changes = today_logs.filter(action__in=['added', 'quantity_taken']).aggregate(
                total_added=db_models.Sum('quantity_change', filter=db_models.Q(quantity_change__gt=0)),
                total_removed=db_models.Sum('quantity_change', filter=db_models.Q(quantity_change__lt=0))
            )
            
            # إعداد البيانات للـ JSON
            report_data = {
                'new_products': [
                    {
                        'product_number': p.product_number,
                        'name': p.name,
                        'category': p.category,
                        'quantity': p.quantity,
                        'location': p.location.full_location if p.location else None
                    } for p in new_products
                ],
                'recent_logs': [
                    {
                        'action': log.action,
                        'product_number': log.product_number,
                        'notes': log.notes,
                        'created_at': log.created_at.strftime('%Y-%m-%d %H:%M')
                    } for log in recent_logs
                ]
            }
            
            # إنشاء التقرير المحفوظ
            hijri_date = convert_to_hijri(today)
            
            DailyReportArchive.objects.create(
                report_date=today,
                hijri_date=hijri_date,
                products_added=products_added,
                products_updated=products_updated,
                products_deleted=products_deleted,
                quantities_taken=quantities_taken,
                locations_assigned=locations_assigned,
                total_added=quantity_changes['total_added'] or 0,
                total_removed=abs(quantity_changes['total_removed'] or 0),
                report_data=report_data
            )
            
            return JsonResponse({
                'success': True,
                'message': 'تم حفظ التقرير بنجاح'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'خطأ: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request'})


def daily_reports_archive(request):
    """عرض السجل (أرشيف التقارير اليومية)"""
    search_query = request.GET.get('search', '')
    
    reports = DailyReportArchive.objects.all()
    
    if search_query:
        # البحث في التاريخ الميلادي والهجري
        reports = reports.filter(
            report_date__icontains=search_query
        ) | reports.filter(
            hijri_date__icontains=search_query
        )
    
    return render(request, 'inventory_app/daily_reports_archive.html', {
        'reports': reports,
        'search_query': search_query
    })


def daily_report_detail(request, report_id):
    """عرض تفاصيل تقرير محفوظ"""
    report = get_object_or_404(DailyReportArchive, id=report_id)
    
    return render(request, 'inventory_app/daily_report_detail.html', {
        'report': report
    })


@exclude_maintenance
@login_required
def backup_restore_page(request):
    """صفحة النسخ الاحتياطي والاستعادة - للمسؤول فقط"""
    # إحصائيات البيانات
    stats = {
        'warehouses': Warehouse.objects.count(),
        'locations': Location.objects.count(),
        'products': Product.objects.count(),
        'audit_logs': AuditLog.objects.count(),
        'daily_reports': DailyReportArchive.objects.count(),
        'orders': Order.objects.count(),
    }
    
    return render(request, 'inventory_app/backup_restore.html', {
        'stats': stats
    })


@csrf_exempt
@require_http_methods(["POST"])
def update_location_notes(request):
    """تحديث ملاحظات موقع"""
    try:
        data = json.loads(request.body)
        location_id = data.get('location_id')
        notes = data.get('notes', '')
        
        if not location_id:
            return JsonResponse({
                'success': False,
                'error': 'معرف الموقع مطلوب'
            })
        
        location = Location.objects.get(id=location_id)
        location.notes = notes
        location.save()
        
        return JsonResponse({
            'success': True,
            'message': 'تم حفظ الملاحظات بنجاح'
        })
        
    except Location.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'الموقع غير موجود'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_http_methods(["POST"])
@exclude_maintenance
@login_required
def export_backup(request):
    """تصدير النسخ الاحتياطي"""
    try:
        # جمع البيانات
        data = {
            'export_info': {
                'date': datetime.now().isoformat(),
                'version': '1.0',
                'description': 'نسخ احتياطي كامل من نظام إدارة المستودع'
            },
            'warehouses': json.loads(serializers.serialize('json', Warehouse.objects.all())),
            'locations': json.loads(serializers.serialize('json', Location.objects.all())),
            'products': json.loads(serializers.serialize('json', Product.objects.all())),
            'audit_logs': json.loads(serializers.serialize('json', AuditLog.objects.all())),
            'daily_reports': json.loads(serializers.serialize('json', DailyReportArchive.objects.all())),
            'orders': json.loads(serializers.serialize('json', Order.objects.all())),
        }
        
        # إنشاء ملف JSON
        json_data = json.dumps(data, ensure_ascii=False, indent=2)
        
        # إعداد الاستجابة
        filename = f'backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        response = HttpResponse(json_data, content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_http_methods(["POST"])
@exclude_maintenance
@login_required
def import_backup(request):
    """استيراد النسخ الاحتياطي"""
    try:
        if 'backup_file' not in request.FILES:
            return JsonResponse({
                'success': False,
                'error': 'لم يتم إرسال ملف'
            })
        
        uploaded_file = request.FILES['backup_file']
        
        # قراءة الملف
        data = json.loads(uploaded_file.read().decode('utf-8'))
        
        # التحقق من صحة الملف
        if 'export_info' not in data:
            return JsonResponse({
                'success': False,
                'error': 'الملف غير صالح'
            })
        
        clear_existing = request.POST.get('clear_existing', 'false') == 'true'
        
        # بدء الاستيراد
        with transaction.atomic():
            if clear_existing:
                # حذف البيانات الموجودة
                AuditLog.objects.all().delete()
                Product.objects.all().delete()
                Location.objects.all().delete()
                Warehouse.objects.all().delete()
                DailyReportArchive.objects.all().delete()
                Order.objects.all().delete()
            
            # استيراد البيانات
            if 'warehouses' in data:
                objects = serializers.deserialize('json', json.dumps(data['warehouses']))
                for obj in objects:
                    obj.save()
            
            if 'locations' in data:
                objects = serializers.deserialize('json', json.dumps(data['locations']))
                for obj in objects:
                    obj.save()
            
            if 'products' in data:
                objects = serializers.deserialize('json', json.dumps(data['products']))
                for obj in objects:
                    obj.save()
            
            if 'audit_logs' in data:
                objects = serializers.deserialize('json', json.dumps(data['audit_logs']))
                for obj in objects:
                    obj.save()
            
            if 'daily_reports' in data:
                objects = serializers.deserialize('json', json.dumps(data['daily_reports']))
                for obj in objects:
                    obj.save()
            
            if 'orders' in data:
                objects = serializers.deserialize('json', json.dumps(data['orders']))
                for obj in objects:
                    obj.save()
        
        return JsonResponse({
            'success': True,
            'message': 'تم الاستيراد بنجاح'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'الملف غير صالح'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@exclude_maintenance
@login_required
def data_deletion_page(request):
    """صفحة حذف البيانات"""
    # إحصائيات البيانات
    stats = {
        'warehouses': Warehouse.objects.count(),
        'locations': Location.objects.count(),
        'products': Product.objects.count(),
        'audit_logs': AuditLog.objects.count(),
        'daily_reports': DailyReportArchive.objects.count(),
        'orders': Order.objects.count(),
    }
    
    return render(request, 'inventory_app/data_deletion.html', {
        'stats': stats
    })


@csrf_exempt
@require_http_methods(["POST"])
@transaction.atomic
@exclude_maintenance
@login_required
def delete_data(request):
    """حذف البيانات المحددة"""
    try:
        data = json.loads(request.body)
        
        # قراءة البيانات المحددة للحذف
        delete_products = data.get('delete_products', False)
        delete_locations = data.get('delete_locations', False)
        delete_warehouses = data.get('delete_warehouses', False)
        delete_audit_logs = data.get('delete_audit_logs', False)
        delete_daily_reports = data.get('delete_daily_reports', False)
        delete_orders = data.get('delete_orders', False)
        
        deleted_items = []
        
        # حذف البيانات المحددة
        if delete_products:
            count = Product.objects.count()
            Product.objects.all().delete()
            deleted_items.append(f'{count} منتج')
        
        if delete_warehouses:
            count = Warehouse.objects.count()
            Warehouse.objects.all().delete()
            deleted_items.append(f'{count} مستودع')
        
        if delete_locations:
            count = Location.objects.count()
            Location.objects.all().delete()
            deleted_items.append(f'{count} مكان')
        
        if delete_audit_logs:
            count = AuditLog.objects.count()
            AuditLog.objects.all().delete()
            deleted_items.append(f'{count} سجل عمليات')
        
        if delete_daily_reports:
            count = DailyReportArchive.objects.count()
            DailyReportArchive.objects.all().delete()
            deleted_items.append(f'{count} تقرير يومي')
        
        if delete_orders:
            count = Order.objects.count()
            Order.objects.all().delete()
            deleted_items.append(f'{count} طلبية')
        
        if not deleted_items:
            return JsonResponse({
                'success': False,
                'error': 'لم يتم تحديد أي بيانات للحذف'
            })
        
        return JsonResponse({
            'success': True,
            'message': f'تم حذف: {", ".join(deleted_items)}'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'البيانات غير صالحة'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


def orders_list(request):
    """عرض قائمة الطلبات المسحوبة"""
    orders = Order.objects.all()
    return render(request, 'inventory_app/orders_list.html', {
        'orders': orders
    })


def order_detail(request, order_id):
    """عرض تفاصيل طلبية محددة"""
    order = get_object_or_404(Order, id=order_id)
    return render(request, 'inventory_app/order_detail.html', {
        'order': order
    })


@csrf_exempt
def delete_order(request, order_id):
    """حذف طلبية محددة"""
    if request.method == 'DELETE':
        try:
            order = get_object_or_404(Order, id=order_id)
            order_number = order.order_number
            order.delete()
            return JsonResponse({
                'success': True,
                'message': f'تم حذف الطلبية {order_number}'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    return JsonResponse({'error': 'Invalid request method'}, status=400)


@login_required
@require_http_methods(["POST"])
@csrf_exempt
@transaction.atomic
def reset_all_quantities(request):
    """تصفير جميع الكميات في قاعدة البيانات"""
    try:
        # جلب جميع المنتجات
        total_products = Product.objects.count()
        
        if total_products == 0:
            return JsonResponse({
                'success': True,
                'message': 'لا توجد منتجات في قاعدة البيانات',
                'count': 0
            })
        
        # تحديث جميع الكميات إلى 0
        updated_count = Product.objects.update(quantity=0)
        
        return JsonResponse({
            'success': True,
            'message': f'تم تصفير الكميات لجميع المنتجات بنجاح ({updated_count} منتج)',
            'count': updated_count
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'حدث خطأ أثناء تصفير الكميات: {str(e)}'
        }, status=500)


# ========== نظام المستخدمين والصلاحيات ==========

@require_http_methods(["GET", "POST"])
def custom_login(request):
    """تسجيل الدخول مخصص مع تسجيل النشاط"""
    if request.user.is_authenticated:
        return redirect('inventory_app:home')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            auth_login(request, user)
            
            # تسجيل نشاط تسجيل الدخول
            UserActivityLog.log_activity(
                user=user,
                action='login',
                description=f'تم تسجيل الدخول بنجاح',
                request=request
            )
            
            # إنشاء UserProfile إذا لم يكن موجوداً
            if not hasattr(user, 'user_profile'):
                UserProfile.objects.create(
                    user=user,
                    user_type='admin' if user.is_superuser else 'staff'
                )
            
            messages.success(request, f'مرحباً {user.username}!')
            return redirect('inventory_app:home')
        else:
            messages.error(request, 'اسم المستخدم أو كلمة المرور غير صحيحة')
            # تسجيل محاولة فاشلة
            if username:
                try:
                    failed_user = User.objects.get(username=username)
                    UserActivityLog.log_activity(
                        user=failed_user,
                        action='login',
                        description=f'محاولة تسجيل دخول فاشلة - كلمة مرور خاطئة',
                        request=request
                    )
                except User.DoesNotExist:
                    pass
            
            return redirect('login')
    
    return render(request, 'auth/login.html')


@login_required
def custom_logout(request):
    """تسجيل الخروج مخصص مع تسجيل النشاط"""
    user = request.user
    
    # تسجيل نشاط تسجيل الخروج
    UserActivityLog.log_activity(
        user=user,
        action='logout',
        description='تم تسجيل الخروج بنجاح',
        request=request
    )
    
    auth_logout(request)
    messages.success(request, 'تم تسجيل الخروج بنجاح')
    return redirect('login')


@admin_required
@require_http_methods(["GET", "POST"])
def register_staff(request):
    """إنشاء حساب موظف جديد - للمسؤول فقط"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        password_confirm = request.POST.get('password_confirm')
        email = request.POST.get('email', '')
        phone = request.POST.get('phone', '')
        
        # التحقق من البيانات
        if not username or not password:
            messages.error(request, 'اسم المستخدم وكلمة المرور مطلوبان')
            return render(request, 'auth/register.html')
        
        if password != password_confirm:
            messages.error(request, 'كلمات المرور غير متطابقة')
            return render(request, 'auth/register.html')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'اسم المستخدم موجود بالفعل')
            return render(request, 'auth/register.html')
        
        try:
            # إنشاء المستخدم
            with transaction.atomic():
                user = User.objects.create_user(
                    username=username,
                    password=password,
                    email=email,
                    is_staff=False  # لا نجعله staff في Django default
                )
                
                # إنشاء UserProfile
                UserProfile.objects.create(
                    user=user,
                    user_type='staff',
                    phone=phone,
                    is_active=True
                )
                
                # تسجيل النشاط
                UserActivityLog.log_activity(
                    user=request.user,
                    action='user_created',
                    description=f'تم إنشاء حساب موظف جديد: {username}',
                    request=request,
                    object_type='User',
                    object_id=user.id,
                    object_name=username
                )
                
                messages.success(request, f'تم إنشاء حساب الموظف {username} بنجاح')
                return redirect('inventory_app:admin_dashboard')
                
        except Exception as e:
            messages.error(request, f'حدث خطأ أثناء إنشاء الحساب: {str(e)}')
            return render(request, 'auth/register.html')
    
    return render(request, 'auth/register.html')


@admin_required
@exclude_admin_dashboard
def admin_dashboard(request):
    """لوحة تحكم المسؤول - عرض تتبع الموظفين"""
    # إحصائيات عامة
    total_staff = UserProfile.objects.filter(user_type='staff', is_active=True).count()
    total_admins = UserProfile.objects.filter(user_type='admin', is_active=True).count()
    
    # إحصائيات الأنشطة (آخر 7 أيام)
    seven_days_ago = timezone.now() - timedelta(days=7)
    recent_activities = UserActivityLog.objects.filter(created_at__gte=seven_days_ago).count()
    today_activities = UserActivityLog.objects.filter(
        created_at__date=timezone.now().date()
    ).count()
    
    # قائمة الموظفين مع إحصائياتهم (جميع الموظفين - نشط وغير نشط)
    staff_members = []
    staff_profiles = UserProfile.objects.filter(user_type='staff').select_related('user')
    
    for profile in staff_profiles:
        user = profile.user
        # عدد الأنشطة اليوم (UserActivityLog)
        today_count = UserActivityLog.objects.filter(
            user=user,
            created_at__date=timezone.now().date()
        ).count()
        
        # عدد الأنشطة آخر 7 أيام (UserActivityLog)
        week_count = UserActivityLog.objects.filter(
            user=user,
            created_at__gte=seven_days_ago
        ).count()
        
        # عدد العمليات اليوم (AuditLog)
        today_operations = AuditLog.objects.filter(
            user=user.username,
            created_at__date=timezone.now().date()
        ).count()
        
        # عدد العمليات آخر 7 أيام (AuditLog)
        week_operations = AuditLog.objects.filter(
            user=user.username,
            created_at__gte=seven_days_ago
        ).count()
        
        # آخر نشاط (UserActivityLog)
        last_activity = UserActivityLog.objects.filter(user=user).order_by('-created_at').first()
        
        # آخر عملية (AuditLog)
        last_operation = AuditLog.objects.filter(user=user.username).order_by('-created_at').first()
        
        # العمليات الأخيرة (آخر 10 عمليات)
        recent_operations = AuditLog.objects.filter(user=user.username).select_related('product').order_by('-created_at')[:10]
        
        # إحصائيات العمليات حسب النوع
        operation_stats = {}
        for action_code, action_name in AuditLog.ACTION_CHOICES:
            count = AuditLog.objects.filter(user=user.username, action=action_code).count()
            if count > 0:
                operation_stats[action_name] = count
        
        staff_members.append({
            'profile': profile,
            'user': user,
            'today_activities': today_count,
            'week_activities': week_count,
            'today_operations': today_operations,
            'week_operations': week_operations,
            'last_activity': last_activity,
            'last_operation': last_operation,
            'recent_operations': recent_operations,
            'operation_stats': operation_stats,
            'last_login_ip': profile.last_login_ip,
        })
    
    # ترتيب الموظفين حسب النشاط
    staff_members.sort(key=lambda x: x['today_activities'], reverse=True)
    
    # آخر 50 نشاط
    recent_logs = UserActivityLog.objects.select_related('user').order_by('-created_at')[:50]
    
    # إحصائيات الأنشطة حسب النوع
    activity_stats = {}
    for action_code, action_name in UserActivityLog.ACTION_TYPES:
        count = UserActivityLog.objects.filter(
            action=action_code,
            created_at__gte=seven_days_ago
        ).count()
        if count > 0:
            activity_stats[action_name] = count
    
    context = {
        'total_staff': total_staff,
        'total_admins': total_admins,
        'recent_activities': recent_activities,
        'today_activities': today_activities,
        'staff_members': staff_members,
        'recent_logs': recent_logs,
        'activity_stats': activity_stats,
    }
    
    # تسجيل النشاط
    UserActivityLog.log_activity(
        user=request.user,
        action='page_viewed',
        description='عرض لوحة تحكم المسؤول',
        request=request,
        url=request.path
    )
    
    return render(request, 'inventory_app/admin_dashboard.html', context)


@staff_required
def staff_dashboard(request):
    """لوحة تحكم الموظف - إحصائيات شخصية"""
    user = request.user
    
    # إحصائيات اليوم
    today = timezone.now().date()
    today_activities = UserActivityLog.objects.filter(
        user=user,
        created_at__date=today
    ).count()
    
    # إحصائيات آخر 7 أيام
    seven_days_ago = timezone.now() - timedelta(days=7)
    week_activities = UserActivityLog.objects.filter(
        user=user,
        created_at__gte=seven_days_ago
    ).count()
    
    # آخر 20 نشاط
    recent_activities = UserActivityLog.objects.filter(user=user).order_by('-created_at')[:20]
    
    # الحصول على UserProfile
    user_profile = None
    if hasattr(user, 'user_profile'):
        user_profile = user.user_profile
    
    context = {
        'user': user,
        'user_profile': user_profile,
        'today_activities': today_activities,
        'week_activities': week_activities,
        'recent_activities': recent_activities,
    }
    
    # تسجيل النشاط
    UserActivityLog.log_activity(
        user=user,
        action='page_viewed',
        description='عرض لوحة تحكم الموظف',
        request=request,
        url=request.path
    )
    
    return render(request, 'inventory_app/staff_dashboard.html', context)


@staff_required
def user_profile(request):
    """الملف الشخصي للمستخدم"""
    user = request.user
    
    # الحصول على UserProfile
    user_profile = None
    if hasattr(user, 'user_profile'):
        user_profile = user.user_profile
    else:
        # إنشاء profile إذا لم يكن موجوداً
        user_profile = UserProfile.objects.create(
            user=user,
            user_type='admin' if user.is_superuser else 'staff'
        )
    
    # جميع الأنشطة
    all_activities = UserActivityLog.objects.filter(user=user).order_by('-created_at')[:100]
    
    # إحصائيات حسب نوع النشاط
    activity_by_type = {}
    for action_code, action_name in UserActivityLog.ACTION_TYPES:
        count = UserActivityLog.objects.filter(user=user, action=action_code).count()
        if count > 0:
            activity_by_type[action_name] = count
    
    context = {
        'user': user,
        'user_profile': user_profile,
        'all_activities': all_activities,
        'activity_by_type': activity_by_type,
    }
    
    # تسجيل النشاط
    UserActivityLog.log_activity(
        user=user,
        action='page_viewed',
        description='عرض الملف الشخصي',
        request=request,
        url=request.path
    )
    
    return render(request, 'inventory_app/user_profile.html', context)


@admin_required
def view_staff(request, user_id):
    """عرض تفاصيل موظف - للمسؤول فقط"""
    staff_user = get_object_or_404(User, id=user_id)
    staff_profile = get_object_or_404(UserProfile, user=staff_user)
    
    # جميع الأنشطة (UserActivityLog)
    all_activities = UserActivityLog.objects.filter(user=staff_user).order_by('-created_at')[:100]
    
    # جميع العمليات المفصلة (AuditLog) - عمليات المنتجات
    all_product_operations = AuditLog.objects.filter(user=staff_user.username).select_related('product').order_by('-created_at')[:200]
    
    # إحصائيات العمليات حسب النوع
    operation_stats = {}
    for action_code, action_name in AuditLog.ACTION_CHOICES:
        count = AuditLog.objects.filter(user=staff_user.username, action=action_code).count()
        if count > 0:
            operation_stats[action_name] = count
    
    # إحصائيات حسب نوع النشاط (UserActivityLog)
    activity_by_type = {}
    for action_code, action_name in UserActivityLog.ACTION_TYPES:
        count = UserActivityLog.objects.filter(user=staff_user, action=action_code).count()
        if count > 0:
            activity_by_type[action_name] = count
    
    # إحصائيات الوقت
    today = timezone.now().date()
    today_activities = UserActivityLog.objects.filter(user=staff_user, created_at__date=today).count()
    today_operations = AuditLog.objects.filter(user=staff_user.username, created_at__date=today).count()
    seven_days_ago = timezone.now() - timedelta(days=7)
    week_activities = UserActivityLog.objects.filter(user=staff_user, created_at__gte=seven_days_ago).count()
    week_operations = AuditLog.objects.filter(user=staff_user.username, created_at__gte=seven_days_ago).count()
    
    context = {
        'staff_user': staff_user,
        'staff_profile': staff_profile,
        'all_activities': all_activities,
        'all_product_operations': all_product_operations,
        'activity_by_type': activity_by_type,
        'operation_stats': operation_stats,
        'today_activities': today_activities,
        'today_operations': today_operations,
        'week_activities': week_activities,
        'week_operations': week_operations,
    }
    
    # تسجيل النشاط
    UserActivityLog.log_activity(
        user=request.user,
        action='user_viewed',
        description=f'عرض ملف الموظف {staff_user.username}',
        request=request,
        object_type='User',
        object_id=staff_user.id,
        object_name=staff_user.username
    )
    
    return render(request, 'inventory_app/view_staff.html', context)


@admin_required
def edit_staff(request, user_id):
    """تعديل بيانات موظف - للمسؤول فقط"""
    staff_user = get_object_or_404(User, id=user_id)
    staff_profile = get_object_or_404(UserProfile, user=staff_user)
    
    if request.method == 'POST':
        # تحديث بيانات المستخدم
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        
        # التحقق من عدم تكرار اسم المستخدم
        if username and username != staff_user.username:
            if User.objects.filter(username=username).exclude(id=user_id).exists():
                messages.error(request, 'اسم المستخدم مستخدم بالفعل')
            else:
                staff_user.username = username
        
        staff_user.email = email
        staff_user.save()
        
        # تحديث UserProfile
        staff_profile.phone = request.POST.get('phone', '').strip()
        staff_profile.notes = request.POST.get('notes', '').strip()
        user_type = request.POST.get('user_type', 'staff')
        if user_type in ['admin', 'staff']:
            staff_profile.user_type = user_type
        staff_profile.save()
        
        # تحديث كلمة السر إذا تم إدخالها
        new_password = request.POST.get('password', '').strip()
        if new_password:
            staff_user.set_password(new_password)
            staff_user.save()
        
        # تسجيل النشاط
        UserActivityLog.log_activity(
            user=request.user,
            action='user_updated',
            description=f'تعديل بيانات الموظف {staff_user.username}',
            request=request,
            object_type='User',
            object_id=staff_user.id,
            object_name=staff_user.username
        )
        
        messages.success(request, f'تم تحديث بيانات الموظف {staff_user.username} بنجاح')
        return redirect('inventory_app:admin_dashboard')
    
    context = {
        'staff_user': staff_user,
        'staff_profile': staff_profile,
    }
    
    return render(request, 'inventory_app/edit_staff.html', context)


@admin_required
@csrf_exempt
@require_http_methods(["POST"])
@transaction.atomic
def toggle_staff_active(request, user_id):
    """تفعيل/تعطيل موظف - للمسؤول فقط"""
    try:
        staff_user = get_object_or_404(User, id=user_id)
        staff_profile = get_object_or_404(UserProfile, user=staff_user)
        
        # منع المسؤول من تعطيل نفسه
        if staff_user == request.user:
            return JsonResponse({
                'success': False,
                'error': 'لا يمكنك تعطيل نفسك'
            }, status=400)
        
        # تبديل الحالة
        staff_profile.is_active = not staff_profile.is_active
        staff_profile.save()
        
        action_text = 'تفعيل' if staff_profile.is_active else 'تعطيل'
        
        # تسجيل النشاط
        UserActivityLog.log_activity(
            user=request.user,
            action='user_updated',
            description=f'{action_text} الموظف {staff_user.username}',
            request=request,
            object_type='User',
            object_id=staff_user.id,
            object_name=staff_user.username,
            details={'is_active': staff_profile.is_active}
        )
        
        return JsonResponse({
            'success': True,
            'message': f'تم {action_text} الموظف {staff_user.username} بنجاح',
            'is_active': staff_profile.is_active
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'حدث خطأ: {str(e)}'
        }, status=500)


@admin_required
@csrf_exempt
@require_http_methods(["POST"])
@transaction.atomic
def delete_staff(request, user_id):
    """حذف موظف - للمسؤول فقط"""
    try:
        staff_user = get_object_or_404(User, id=user_id)
        staff_profile = get_object_or_404(UserProfile, user=staff_user)
        
        # منع المسؤول من حذف نفسه
        if staff_user == request.user:
            return JsonResponse({
                'success': False,
                'error': 'لا يمكنك حذف نفسك'
            }, status=400)
        
        username = staff_user.username
        
        # تسجيل النشاط قبل الحذف
        UserActivityLog.log_activity(
            user=request.user,
            action='user_deleted',
            description=f'حذف الموظف {username}',
            request=request,
            object_type='User',
            object_id=staff_user.id,
            object_name=username
        )
        
        # حذف UserProfile أولاً
        staff_profile.delete()
        # حذف User
        staff_user.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'تم حذف الموظف {username} بنجاح'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'حدث خطأ: {str(e)}'
        }, status=500)

